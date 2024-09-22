import pymysql
import pandas as pd
import main
from datetime import datetime
import os

# 경로 설정
db_value_list3 = main.result_list3
db_value_list4 = main.result_list4
db_value_list = main.result_list
num_color = ""

db = pymysql.connect(host="localhost", user="root", password="0000", charset="utf8")
current_datetime = datetime.now()
current_date = current_datetime.date()

days_of_week = ['월', '화', '수', '목', '금', '토', '일']
current_day_of_week = days_of_week[current_date.weekday()]
output_dir1 = "C:\\subway_excel_data"
output_dir = os.path.join(output_dir1, current_day_of_week)
os.makedirs(output_dir, exist_ok=True)  # 디렉터리가 없으면 생성

cursor = db.cursor()

# 데이터베이스 선택
cursor.execute("USE testdb")

insert_query = "INSERT INTO traffic_data (date, day_of_week, time, area, color) VALUES (%s, %s, %s, %s, %s)"
for db_value in db_value_list:
    if 11 <= db_value:
        num_color = "초록"
    
    elif 9 < db_value < 11:
        num_color = "노랑"
        
    else:
        num_color = "빨강"
    
    data_to_insert = (current_date.strftime('%Y-%m-%d'), current_day_of_week, current_datetime.strftime('%H:%M:%S'), db_value, num_color)

    cursor.execute(insert_query, data_to_insert)
    db.commit()
cursor.execute("USE testdb")

insert_query2 = "INSERT INTO traffic_data_csv (date, day_of_week, time, area, color) VALUES (%s, %s, %s, %s, %s)"
current_time = datetime.now().time()
time_in_minutes = current_time.hour * 100 + current_time.minute

for i in range(len(db_value_list)):
    db_value = db_value_list[i]
    
    if 11 <= db_value:
        color_num = 2
    elif 9 < db_value < 11:
        color_num = 1
    else:
        color_num = 3

    if current_day_of_week == "월":
        data_num = 1
    elif current_day_of_week == "화":
        data_num = 2
    elif current_day_of_week == "수":
        data_num = 3
    elif current_day_of_week == "목":
        data_num = 4
    elif current_day_of_week == "금":
        data_num = 5
    elif current_day_of_week == "토":
        data_num = 6
    else:
        data_num = 7
    
    data_to_insert2 = (current_date.strftime('%Y-%m-%d'), data_num, time_in_minutes, db_value, color_num)

    cursor.execute(insert_query2, data_to_insert2)
db.commit()






# 엑셀 파일 생성
query = "SELECT * FROM traffic_data"
df = pd.read_sql_query(query, db)
df['time'] = df['time'].astype(str).str.replace('0 days ', '')
df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
file_name = "subway_excel_data_" + current_date.strftime('%Y-%m-%d') + ".xlsx"
file_path = os.path.join(output_dir, file_name)
df.to_excel(file_path, index=False)


# 엑셀 셀 배경색 변경
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 엑셀 파일 열기
wb = load_workbook(file_path)
ws = wb.active

# 조건에 따라 area 열의 배경색 설정
for row_num, area_value in enumerate(df['area'], start=2):
    cell = ws.cell(row=row_num, column=5)

    if 10.5 < area_value:
        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 초록 배경색
    elif 8.0 < area_value <= 10.5:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 노랑 배경색
    else:
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 빨강 배경색

# 엑셀 파일 저장
wb.save(file_path)

query2 = "SELECT * FROM traffic_data_csv"
dd = pd.read_sql_query(query2, db)
dd['time'] = dd['time'].astype(str).str.replace('0 days ', '')
dd['date'] = pd.to_datetime(dd['date']).dt.strftime('%Y-%m-%d')
file_name_scv = "subway_excel_data_" + current_date.strftime('%Y-%m-%d') + ".csv"
file_path_csv = os.path.join(output_dir, file_name_scv)
dd.to_csv(file_path_csv, index=False)

db.close()
